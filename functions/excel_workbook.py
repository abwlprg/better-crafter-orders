"""Excel workbook helpers for supplier-specific order targets.

These helpers operate on local XLSX bytes only. They never call OneDrive,
delete rows, or move existing worksheets.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Iterable

from openpyxl import Workbook, load_workbook


BASE_COLUMNS = [
    "Date",
    "Item No.",
    "QTY",
    "Color",
    "Customer Name",
    "Sent to Supplier",
    "Ship by date",
    "Sent to customer",
]


def normalize_column_name(name: str | None) -> str:
    return " ".join(str(name or "").lower().strip().split())


@dataclass(slots=True)
class WorkbookWriteResult:
    content: bytes
    rows_written: int
    duplicates_skipped: int
    columns_added: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def create_workbook_with_sheet(sheet_name: str, headers: Iterable[str] | None = None) -> bytes:
    """Create a workbook with one sheet and a header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = str(sheet_name)
    ws.append(list(headers or BASE_COLUMNS))
    return _save(wb)


def create_year_sheet(
    xlsx_bytes: bytes,
    sheet_name: str,
    source_sheet: str | None = None,
) -> tuple[bytes, bool]:
    """Create a yearly sheet by copying only headers from an existing sheet.

    Returns (content, created). If the sheet already exists, the original bytes
    are returned and created is False.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    name = str(sheet_name)
    if name in wb.sheetnames:
        return xlsx_bytes, False

    source = wb[source_sheet] if source_sheet and source_sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [cell.value or "" for cell in source[1]]
    ws = wb.create_sheet(name)
    ws.append(headers)
    return _save(wb), True


def add_missing_columns(
    xlsx_bytes: bytes,
    sheet_name: str,
    required_columns: Iterable[str],
) -> tuple[bytes, list[str]]:
    """Add missing header columns to a sheet without touching existing rows."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' does not exist")
    ws = wb[sheet_name]
    existing = [str(cell.value or "").strip() for cell in ws[1]]
    existing_normalized = {normalize_column_name(name) for name in existing if name}
    added: list[str] = []
    for col in required_columns:
        name = str(col).strip()
        normalized = normalize_column_name(name)
        if name and normalized not in existing_normalized:
            ws.cell(row=1, column=len(existing) + len(added) + 1, value=name)
            added.append(name)
            existing_normalized.add(normalized)
    if not added:
        return xlsx_bytes, []
    return _save(wb), added


def append_orders(
    xlsx_bytes: bytes,
    sheet_name: str,
    orders: list[dict],
    custom_fields: list[dict] | None = None,
    seen_keys: set[str] | None = None,
) -> WorkbookWriteResult:
    """Append orders to the selected sheet, preserving all existing data."""
    from functions.idempotency import build_idempotency_key

    custom_columns = [
        str(field.get("field_name", "")).strip()
        for field in custom_fields or []
        if isinstance(field, dict) and str(field.get("field_name", "")).strip()
    ]
    required_columns = BASE_COLUMNS + custom_columns
    content, added = add_missing_columns(xlsx_bytes, sheet_name, required_columns)

    wb = load_workbook(io.BytesIO(content))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' does not exist")
    ws = wb[sheet_name]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    header_index = {normalize_column_name(name): idx + 1 for idx, name in enumerate(headers)}
    seen = set(seen_keys or set())
    warnings: list[str] = []
    written = 0
    duplicates = 0

    for order in orders:
        try:
            key = build_idempotency_key(order)
        except ValueError as exc:
            warnings.append(f"Skipped row without idempotency key: {exc}")
            continue
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        row = ws.max_row + 1
        values = {
            "Date": order.get("order_date", ""),
            "Item No.": order.get("item_code", ""),
            "QTY": order.get("quantity", ""),
            "Color": order.get("color", ""),
            "Customer Name": order.get("customer_name", ""),
            "Sent to Supplier": "y",
            "Ship by date": order.get("ship_by", ""),
            "Sent to customer": "",
        }
        for col in custom_columns:
            values[col] = order.get(col, "")
        for header, value in values.items():
            normalized_header = normalize_column_name(header)
            if normalized_header in header_index:
                ws.cell(row=row, column=header_index[normalized_header], value=value)
        written += 1

    return WorkbookWriteResult(
        content=_save(wb),
        rows_written=written,
        duplicates_skipped=duplicates,
        columns_added=added,
        warnings=warnings,
    )


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
