"""Tests for supplier Excel workbook helpers using test-only data."""

from __future__ import annotations

import io
import unittest

from openpyxl import load_workbook

from functions.excel_workbook import (
    BASE_COLUMNS,
    add_missing_columns,
    append_orders,
    create_workbook_with_sheet,
    create_year_sheet,
)


def _load(content: bytes):
    return load_workbook(io.BytesIO(content))


class TestSupplierExcelWorkbook(unittest.TestCase):
    def test_creating_new_year_sheet_copies_headers_only(self) -> None:
        content = create_workbook_with_sheet("2026", BASE_COLUMNS)
        wb = _load(content)
        ws = wb["2026"]
        ws.append(["06/02/2026", "JAKE-TEST-1", "1", "", "Jake Test"])
        buf = io.BytesIO()
        wb.save(buf)

        updated, created = create_year_sheet(buf.getvalue(), "2027", source_sheet="2026")

        self.assertTrue(created)
        wb2 = _load(updated)
        self.assertEqual([c.value for c in wb2["2027"][1]], BASE_COLUMNS)
        self.assertEqual(wb2["2027"].max_row, 1)
        self.assertEqual(wb2["2026"].max_row, 2)

    def test_existing_year_sheet_is_not_overwritten(self) -> None:
        content = create_workbook_with_sheet("2026", BASE_COLUMNS)

        updated, created = create_year_sheet(content, "2026")

        self.assertFalse(created)
        self.assertEqual(updated, content)

    def test_custom_fields_create_columns_safely_and_preserve_rows(self) -> None:
        content = create_workbook_with_sheet("2026", BASE_COLUMNS)
        wb = _load(content)
        wb["2026"].append(["06/02/2026", "HARVEY-TEST-1", "1", "", "Harvey Test"])
        buf = io.BytesIO()
        wb.save(buf)

        updated, added = add_missing_columns(buf.getvalue(), "2026", BASE_COLUMNS + ["Test Finish"])

        self.assertEqual(added, ["Test Finish"])
        wb2 = _load(updated)
        self.assertEqual(wb2["2026"].max_row, 2)
        self.assertEqual(wb2["2026"]["B2"].value, "HARVEY-TEST-1")
        self.assertIn("Test Finish", [c.value for c in wb2["2026"][1]])

    def test_existing_custom_column_match_is_normalized(self) -> None:
        content = create_workbook_with_sheet("2026", BASE_COLUMNS + ["Order Number"])

        updated, added = add_missing_columns(content, "2026", BASE_COLUMNS + [" order   number "])

        self.assertEqual(added, [])
        wb = _load(updated)
        headers = [c.value for c in wb["2026"][1]]
        self.assertEqual(headers.count("Order Number"), 1)

    def test_rows_write_only_to_assigned_active_sheet_and_dedupe(self) -> None:
        content = create_workbook_with_sheet("2026", BASE_COLUMNS)
        content, _ = create_year_sheet(content, "2027", source_sheet="2026")
        order = {
            "supplier_id": "jake-test",
            "thread_id": "thread-jake-test",
            "message_id": "msg-jake-test",
            "item_index": "1",
            "order_date": "2026-06-02",
            "item_code": "JAKE-TEST-2",
            "quantity": "3",
            "customer_name": "Jake Test",
        }

        result = append_orders(content, "2027", [order, dict(order)])

        self.assertEqual(result.rows_written, 1)
        self.assertEqual(result.duplicates_skipped, 1)
        wb = _load(result.content)
        self.assertEqual(wb["2026"].max_row, 1)
        self.assertEqual(wb["2027"].max_row, 2)
        self.assertEqual(wb["2027"]["B2"].value, "JAKE-TEST-2")


if __name__ == "__main__":
    unittest.main()
